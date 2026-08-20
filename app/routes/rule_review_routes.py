"""Rule Validation tab.

Page:
  GET  /rule-review

API (all read-only against FortiManager; POST is for submitting work items):
  GET  /api/rule-review/adoms
  GET  /api/rule-review/adoms/<adom>/packages
  POST /api/rule-review/parse-import        — parse uploaded CSV or XLSX
  POST /api/rule-review/analyze             — run analysis
  GET  /api/rule-review/zone-status         — is zone policy DB available?
"""

import csv
import io

from flask import Blueprint, current_app, jsonify, render_template, request, session
from app.decorators import tab_required, check_adom_access
from app.fmg_helpers import make_client
from app.fmg_client import FMGError
from app import registry
from app.rule_review import analyze_flows, zone_script_available
from app.security import internal_api_error, upstream_api_error

bp = Blueprint("rule_review", __name__)

registry.register("rule_review", "Rule Validation", "rule_review.rule_review_page")


# ── Page ──────────────────────────────────────────────────────────────────────


@bp.route("/rule-review")
@tab_required("rule_review")
def rule_review_page():
    return render_template("rule_review.html", user=session["user"])


# ── API: ADOM list ────────────────────────────────────────────────────────────


@bp.route("/api/rule-review/adoms")
@tab_required("rule_review")
def rr_adoms():
    try:
        from flask import session as _session
        from app.groups import get_allowed_adoms

        allowed = get_allowed_adoms(
            _session.get("user", ""), ad_groups=_session.get("ad_groups", [])
        )
        with make_client() as client:
            raw = client.get_adoms()
        names = sorted(
            a["name"]
            for a in raw
            if isinstance(a, dict)
            and a.get("name")
            and not a["name"].lower().startswith("forti")
        )
        if allowed is not None:
            names = [n for n in names if n in allowed]
        return jsonify(names)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)


# ── API: package list ─────────────────────────────────────────────────────────


@bp.route("/api/rule-review/adoms/<adom>/packages")
@tab_required("rule_review")
def rr_packages(adom: str):
    if err := check_adom_access(adom):
        return err
    try:
        with make_client() as client:
            raw = client.get_policy_packages(adom)
        packages = [
            {"name": p["name"], "path": p.get("path", p["name"])}
            for p in raw
            if isinstance(p, dict)
            and p.get("name")
            and (p.get("type") or "").lower() != "folder"
        ]
        return jsonify(packages)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)


# ── API: device list ──────────────────────────────────────────────────────────


@bp.route("/api/rule-review/adoms/<adom>/devices")
@tab_required("rule_review")
def rr_devices(adom: str):
    if err := check_adom_access(adom):
        return err
    try:
        with make_client() as client:
            raw = client.get_devices(adom)
        devices = sorted(
            [
                {"name": d["name"], "ip": d.get("ip", "")}
                for d in raw
                if isinstance(d, dict) and d.get("name")
            ],
            key=lambda d: d["name"].lower(),
        )
        return jsonify(devices)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception:
        return jsonify([])


# ── API: VDOM list ────────────────────────────────────────────────────────────


@bp.route("/api/rule-review/adoms/<adom>/devices/<device>/vdoms")
@tab_required("rule_review")
def rr_device_vdoms(adom: str, device: str):
    if err := check_adom_access(adom):
        return err
    try:
        with make_client() as client:
            raw = client.get_device_vdoms(adom, device)
        names = sorted(
            {d["name"] for d in raw if isinstance(d, dict) and d.get("name")},
            key=str.lower,
        )
        if not names or names == ["root"]:
            return jsonify(["root"])
        return jsonify(names)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception:
        return jsonify(["root"])


# ── API: parse import file ────────────────────────────────────────────────────


@bp.route("/api/rule-review/parse-import", methods=["POST"])
@tab_required("rule_review")
def rr_parse_import():
    """Accept a CSV or XLSX upload, return parsed rows.

    Expected columns (case-insensitive, flexible order):
      src / source / source_ip
      dst / destination / dest / destination_ip / dest_ip
      port / service / ports / services
      comment / comments / note / notes  (optional)
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    filename = (f.filename or "").lower()
    mimetype = (f.mimetype or "").lower()

    allowed_csv_types = {
        "text/csv",
        "application/csv",
        "application/vnd.ms-excel",
        "application/octet-stream",
    }
    allowed_xlsx_types = {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    }

    def _file_size(stream) -> int:
        pos = stream.tell()
        stream.seek(0, 2)
        size = stream.tell()
        stream.seek(pos)
        return size

    max_bytes = int(current_app.config.get("MAX_CONTENT_LENGTH", 4 * 1024 * 1024))
    if _file_size(f.stream) > max_bytes:
        return jsonify({"error": "Uploaded file is too large"}), 413

    rows: list[dict] = []
    errors: list[str] = []

    if filename.endswith(".csv"):
        if mimetype not in allowed_csv_types:
            return jsonify({"error": "Unsupported CSV content type"}), 400
        try:
            content = f.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            rows, errors = _parse_rows(reader)
        except Exception:
            return jsonify({"error": "CSV parse error"}), 400

    elif filename.endswith((".xls", ".xlsx")):
        if mimetype not in allowed_xlsx_types:
            return jsonify({"error": "Unsupported XLSX content type"}), 400
        try:
            import openpyxl

            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(c.value or "").strip().lower()
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            dict_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                dict_rows.append(
                    {
                        headers[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row)
                    }
                )
            rows, errors = _parse_rows(dict_rows)
        except ImportError:
            return jsonify(
                {"error": "openpyxl not installed — XLSX import unavailable"}
            ), 500
        except Exception:
            return jsonify({"error": "XLSX parse error"}), 400
    else:
        return jsonify(
            {"error": "Unsupported file type. Upload a .csv or .xlsx file."}
        ), 400

    return jsonify({"rows": rows, "errors": errors})


_SRC_ALIASES = {"src", "source", "source_ip", "src_ip"}
_DST_ALIASES = {"dst", "destination", "dest", "destination_ip", "dest_ip", "dst_ip"}
_SVC_ALIASES = {"port", "ports", "service", "services", "svc"}
_CMT_ALIASES = {"comment", "comments", "note", "notes"}


def _canonical_header(raw: str) -> str:
    key = raw.strip().lower().replace(" ", "_").replace("-", "_")
    if key in _SRC_ALIASES:
        return "src"
    if key in _DST_ALIASES:
        return "dst"
    if key in _SVC_ALIASES:
        return "service"
    if key in _CMT_ALIASES:
        return "comment"
    return key


def _parse_rows(reader) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    errors: list[str] = []
    for line_no, raw_row in enumerate(reader, start=2):
        row = {
            _canonical_header(k): (v or "").strip()
            for k, v in (raw_row.items() if hasattr(raw_row, "items") else {}.items())
        }
        src = row.get("src", "")
        dst = row.get("dst", "")
        svc = row.get("service", "")
        if not src and not dst:
            continue  # skip blank rows
        if not src:
            errors.append(f"Row {line_no}: missing source IP")
            continue
        if not dst:
            errors.append(f"Row {line_no}: missing destination IP")
            continue
        rows.append(
            {"src": src, "dst": dst, "service": svc, "comment": row.get("comment", "")}
        )
    return rows, errors


# ── API: zone policy status ───────────────────────────────────────────────────


@bp.route("/api/rule-review/zone-status")
@tab_required("rule_review")
def rr_zone_status():
    return jsonify({"available": zone_script_available()})


# ── API: analyze ─────────────────────────────────────────────────────────────


@bp.route("/api/rule-review/analyze", methods=["POST"])
@tab_required("rule_review")
def rr_analyze():
    """Run the policy review analysis.

    Request body::
        {
            "flows": [{"src": "...", "dst": "...", "service": "...", "comment": "..."}, ...],
            "selections": [{"adom": "...", "device": "...", "vdoms": ["root", ...]}, ...],
            "metadata": {"change_number": "...", "owner": "...", "justification": "..."}
        }
    """
    data = request.get_json(silent=True) or {}
    flows = data.get("flows", [])
    selections = data.get("selections")
    metadata = data.get("metadata", {})

    if not flows:
        return jsonify({"error": "No flows provided"}), 400
    if selections is None:
        return jsonify(
            {"error": "No selections provided — use 'selections' key (not 'packages')"}
        ), 400
    if not selections:
        return jsonify({"error": "No firewalls selected"}), 400

    if not isinstance(selections, list) or not all(
        isinstance(s, dict) for s in selections
    ):
        return jsonify(
            {
                "error": "selections must be a list of objects with adom, device, and vdoms fields"
            }
        ), 400

    for s in selections:
        if not s.get("adom") or not s.get("device"):
            return jsonify(
                {
                    "error": "each selection must have non-empty 'adom' and 'device' fields"
                }
            ), 400

    # Collect unique ADOMs
    adoms = list(dict.fromkeys(s["adom"] for s in selections if s.get("adom")))

    for adom in adoms:
        if err := check_adom_access(adom):
            return err

    # Pre-collected error results for device/VDOM with no installed package
    error_results: list[dict] = []

    try:
        with make_client() as client:
            # Resolve selections → packages using package scope member data
            packages: list[dict] = []
            pkg_cache: dict[
                str, list
            ] = {}  # adom → all packages (avoid redundant calls)

            for sel in selections:
                adom = sel["adom"]
                device = sel["device"]
                if adom not in pkg_cache:
                    pkg_cache[adom] = client.get_policy_packages(adom)
                all_pkgs = pkg_cache[adom]

                for vdom in sel.get("vdoms") or ["root"]:
                    matched_pkg = None
                    for pkg in all_pkgs:
                        scope = pkg.get("scope member") or pkg.get("scope_member") or []
                        for m in scope:
                            if (
                                isinstance(m, dict)
                                and m.get("name", "").lower() == device.lower()
                                and (not m.get("vdom") or m.get("vdom") == vdom)
                            ):
                                matched_pkg = pkg
                                break
                        if matched_pkg:
                            break

                    if matched_pkg:
                        packages.append(
                            {
                                "adom": adom,
                                "name": matched_pkg["name"],
                                "path": matched_pkg.get("path", matched_pkg["name"]),
                                "device": device,
                                "vdom": vdom,
                            }
                        )
                    else:
                        msg = (
                            f"No policy package found for {device} / VDOM: {vdom}"
                            f' in ADOM "{adom}"'
                        )
                        for flow in flows:
                            error_results.append(
                                {
                                    "src": flow.get("src", ""),
                                    "dst": flow.get("dst", ""),
                                    "service": flow.get("service", ""),
                                    "adom": adom,
                                    "pkg_name": f"{device}/{vdom}",
                                    "pkg_path": "",
                                    "device": device,
                                    "vdom": vdom,
                                    "verdict": "ERROR",
                                    "error": msg,
                                    "matching_rules": [],
                                    "modifiable_rules": [],
                                    "notes": [msg],
                                    "fortios_cli": "",
                                    "object_plans": [],
                                    "approval": {},
                                    "alternative": None,
                                    "permissiveness_warnings": [],
                                    "zone_available": False,
                                    "zone_verdict": "UNAVAILABLE",
                                    "zone_src": [],
                                    "zone_dst": [],
                                    "zone_governing": [],
                                    "zone_all_policies": [],
                                    "path_in_path": None,
                                    "path_confidence": "low",
                                    "path_src_iface": None,
                                    "path_dst_iface": None,
                                    "path_src_route": None,
                                    "path_dst_route": None,
                                    "path_notes": [],
                                    "path_src_reachable": False,
                                    "path_dst_reachable": False,
                                }
                            )

            # Fetch policies for each resolved package
            policies_by_pkg: dict[str, list] = {}
            for pkg in packages:
                adom = pkg["adom"]
                path = pkg["path"]
                key = f"{adom}/{path}"
                try:
                    policies_by_pkg[key] = client.get_policies(adom, path)
                except Exception:
                    policies_by_pkg[key] = []

            # Fetch address and service objects per ADOM
            addr_objects: list = []
            addr_groups: list = []
            svc_objects: list = []
            svc_groups: list = []
            seen_adoms: set[str] = set()
            for adom in adoms:
                if adom in seen_adoms:
                    continue
                seen_adoms.add(adom)
                addr_objects.extend(client.get_address_objects(adom))
                addr_groups.extend(client.get_address_groups(adom))
                svc_objects.extend(client.get_service_objects(adom))
                svc_groups.extend(client.get_service_groups(adom))

            # Fetch routing + interface data
            routing_by_device: dict[str, dict] = {}
            for pkg in packages:
                adom = pkg["adom"]
                path = pkg["path"]
                device = pkg.get("device", "")
                scope = client.get_pkg_scope_members(adom, path)
                device_names = (
                    [m.get("name", m) if isinstance(m, dict) else str(m) for m in scope]
                    if scope
                    else []
                )
                if not device_names and device:
                    device_names = [device]
                for dev_name in device_names:
                    if dev_name in routing_by_device:
                        continue
                    try:
                        ifaces = client.get_device_interfaces_all_vdoms(adom, dev_name)
                        routes = client.get_device_routes_all_vdoms(adom, dev_name)
                        routing_by_device[dev_name] = {
                            "interfaces": ifaces,
                            "routes": routes,
                        }
                        if not pkg.get("device"):
                            pkg["device"] = dev_name
                    except Exception:
                        routing_by_device[dev_name] = {"interfaces": [], "routes": []}

    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)

    analysis_results = []
    if packages:
        analysis_results = analyze_flows(
            requested_flows=flows,
            packages=packages,
            policies_by_pkg=policies_by_pkg,
            addr_objects=addr_objects,
            addr_groups=addr_groups,
            svc_objects=svc_objects,
            svc_groups=svc_groups,
            routing_by_device=routing_by_device,
        )

    all_results = error_results + analysis_results

    return jsonify(
        {
            "results": all_results,
            "metadata": metadata,
            "zone_available": zone_script_available(),
        }
    )
